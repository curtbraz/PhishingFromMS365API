import datetime as dt
import re
import uuid
from O365 import Account

# 1) Login at Azure Portal (App Registrations) @ https://portal.azure.com/#blade/Microsoft_AAD_RegisteredApps/ApplicationsListBlade

# 2) Create an app. Set a name.

# 3) In Supported account types choose "Accounts in any organizational directory and personal Microsoft accounts (e.g. Skype, Xbox, Outlook.com)", if you are using a personal account.

# 4) Set the redirect uri (Web) to: https://login.microsoftonline.com/common/oauth2/nativeclient and click register. This needs to be inserted into the "Redirect URI" text box as simply checking the check box next to this link seems to be insufficent. This is the default redirect uri used by this library, but you can use any other if you want.

# 5) Write down the Application (client) ID. You will need this value.

# 6) Under "Certificates & secrets", generate a new client secret. Set the expiration preferably to never. Write down the value of the client secret created now. It will be hidden later on.

# 7) Under Api Permissions:

    #When authenticating "on behalf of a user":
            #1. Add the delegated permissions for Microsoft Graph you want (see scopes).
            
            #2. It is highly recommended to add "offline_access" permission. If not the user you will have to re-authenticate every hour.
            
    #When authenticating "with your own identity":
            #1. Add the application permissions for Microsoft Graph you want.
            
            #2. Click on the Grant Admin Consent button (if you have admin permissions) or wait until the admin has given consent to your application.
            
# 8) To read and send emails use:

        #Microsoft Graph - Mail - Mail.ReadWrite
        #Microsoft Graph - Mail - Mail.Send
        #Microsoft Graph - User - User.Read

# 9) Then you need to login for the first time to get the access token that will grant access to the user resources.

## ENTER YOUR APPLICATION ID AND SECRET KEY BELOW
credentials = ('74f98c5b-cdc9-.....', 'J-a8Q~fdW7......')

account = Account(credentials)
if account.authenticate(scopes=['basic', 'message_all']):
   print('Authenticated!')

from openpyxl import load_workbook

rowselector = 2

## THIS IS THE NAME/PATH OF YOUR EXCEL FILE
##fn = 'CurtisTestList.xlsx'
fn = 'AllXOEmails.xlsx'

## CHANGE THIS TO THE SHEET IN QUESTION
wb = load_workbook(filename = fn)
sheet = wb['TargetList']

row_count = sheet.max_row

# Function for Creating Calendar Event
def send_email(account,email,row,name,dept,div,tit):
    m = account.new_message()
    # This Automatically Adds the Email from the Spreadsheet as the Recipient
    m.to.add(email)
    ## Subject of the Email
    m.subject = dept + ' Team!'
    
    ## Filename/Path of HTML Template for Email Body
    with open('GiftCard.html', 'r') as file:
        data = file.read().replace('\n', '')

    ## REGEX to Find and Replace HTML Value with Spreadsheet Name. Edit your HTML Template to include <STARTNAME><ENDNAME> Anywhere you Want it Replaced
    newdata = re.sub('<STARTNAME>.*?<ENDNAME>',name,data, flags=re.DOTALL)
    ## URL to Fake Landing Page
    url = 'https://auth.xtenops.com/u/login/?redir=xogift&id=' + str(row)
    ## Path to Tracking Pixel (Dynamic Content)
   # trackingurl = 'https://.../emailopens.php/?id=' + str(row)
    ## REGEX to Find and Replace URL Value in HTML Template, Like the Name Field Above. Change in Template to Match.
    newdata2 = re.sub('<STARTURL>.*?<ENDURL>',url,newdata, flags=re.DOTALL)
    ## REGEX to Find and Replace Tracking Pixel URL Value in HTML Template, Like the Name Field Above. Change in Template to Match.
    #newdata3 = re.sub('<TRACKSTART>.*?<TRACKEND>',trackingurl,newdata2, flags=re.DOTALL)
    newdata3 = re.sub('<STARTDIVISION>.*?<ENDDIVISION>',div,newdata2, flags=re.DOTALL)
    newdata4 = re.sub('<STARTDEPARTMENT>.*?<ENDDEPARTMENT>',dept,newdata3, flags=re.DOTALL)
    newdata5 = re.sub('<STARTTITLE>.*?<ENDTITLE>',tit,newdata4, flags=re.DOTALL)
    
    m.body = newdata5 
    m.send()
    
    # Prints to Terminal
    print('Sending an email to ' + email)
    
    return

# Loops Through Each Row of the Spreadsheet
while rowselector <= row_count:
  from datetime import datetime
  now = datetime.now()
  ## Update With Column of Email Addresses
  cur_cell = 'C' + str(rowselector)
  ## Blank Column for Updating Time Sent
  update_cell = 'I' + str(rowselector)
  ## Blank Column for Updating UUID
  id_cell = 'A' + str(rowselector)
  ## Update With Column of Name of Recipient
  name_cell = 'D' + str(rowselector)
  ## Update With Column of Department
  department_cell = 'G' + str(rowselector)
  ## Update With Column of Division
  division_cell = 'H' + str(rowselector)
  ## Update With Column of Title
  title_cell = 'F' + str(rowselector)
  #uniqueid = str(uuid.uuid4())
  #sheet[id_cell] = uniqueid
  email = sheet[cur_cell].value
  namevalue = sheet[name_cell].value
  uniqueid = sheet[id_cell].value
  department = sheet[department_cell].value
  division = sheet[division_cell].value
  title = sheet[title_cell].value
  send_email(account,email,uniqueid,namevalue,department,division,title)
  sheet[update_cell] = now.strftime("%m/%d/%Y, %H:%M:%S")
  wb.save(fn)
  rowselector += 1
