import datetime as dt
import uuid
from O365 import Account

# 1) Login at Azure Portal (App Registrations) @ https://portal.azure.com/#blade/Microsoft_AAD_RegisteredApps/ApplicationsListBlade

# 2) Create an app. Set a name.

# 3) In Supported account types choose "Accounts in any organizational directory and personal Microsoft accounts (e.g. Skype, Xbox, Outlook.com)", if you are using a personal account.

# 4) Set the redirect uri (Web) to: https://login.microsoftonline.com/common/oauth2/nativeclient and click register. This needs to be inserted into the "Redirect URI" text box as simply checking the check box next to this link seems to be insufficent. This is the default redirect uri used by this library, but you can use any other if you want.

# 5) Write down the Application (client) ID. You will need this value.

# 6) Under "Certificates & secrets", generate a new client secret. Set the expiration preferably to never. Write down the value of the client secret created now. It will be hidden later on.

# 7) Under Api Permissions:

    #When authenticating "on behalf of a user":
            #1. Add the delegated permissions for Microsoft Graph you want (see scopes).
            
            #2. It is highly recommended to add "offline_access" permission. If not the user you will have to re-authenticate every hour.
            
    #When authenticating "with your own identity":
            #1. Add the application permissions for Microsoft Graph you want.
            
            #2. Click on the Grant Admin Consent button (if you have admin permissions) or wait until the admin has given consent to your application.
            
# 8) To read and send emails use:

        #Microsoft Graph - Calendars - Calendars.ReadWrite
        #Microsoft Graph - Mail - Mail.Send
        #Microsoft Graph - User - User.Read

# 9) Then you need to login for the first time to get the access token that will grant access to the user resources.

## ENTER YOUR APPLICATION ID AND SECRET KEY BELOW
credentials = ('APPLICATION/CLIENT ID', 'SECRET')

# the default protocol will be Microsoft Graph
# the default authentication method will be "on behalf of a user"

account = Account(credentials)
if account.authenticate(scopes=['basic', 'message_all']):
   print('Authenticated!')

from openpyxl import load_workbook

rowselector = 2

## THIS IS THE NAME/PATH OF YOUR EXCEL FILE
fn = 'Template.xlsx'

## CHANGE THIS TO THE SHEET IN QUESTION
wb = load_workbook(filename = fn)
sheet = wb['Sheet1']

row_count = sheet.max_row

# Function for Creating Calendar Event
def insert_cal_event(account,email,row):
    schedule = account.schedule()

    calendar = schedule.get_default_calendar()
    new_event = calendar.new_event()  # creates a new unsaved event 
    
    ## Calendar Event Subject
    new_event.subject = 'Mandatory Zoom Meeting'
    ## Calendar Event Location
    new_event.location = 'Zoom Meeting'
    # Adds Recipient from Spreadsheet
    new_event.attendees.add(email)
    # Body of Calender Event (Inserts Unique ID Created in Spreadsheet for Tracking Clicks)
    new_event.body = 'Mandatory all-hands staff meeting to review important new updates regarding the Work From Home Policy. View more details by joining with the Zoom client at https://..../?id=' + str(row) + '.'

    from datetime import datetime, timedelta
    
    def ceil_dt(dt, delta):
        return dt + (datetime.min - dt) % delta

    now = datetime.now()
    
    # Meeting Start Time from Now
    d1 = ceil_dt(now, timedelta(minutes=60))
    
    new_event.start = d1
    
    # Reminder in Minutes Before Meeting Starts
    new_event.remind_before_minutes = 1
    new_event.save()
    
    # Print to Terminal
    print('Sending an invite to ' + email + ' for ' + new_event.start.strftime("%m/%d/%Y %H:%M:%S"))
    
    return

# Loops Through Each Row of the Spreadsheet
while rowselector <= row_count:
  from datetime import datetime
  now = datetime.now()
  ## Update With Column of Email Addresses
  cur_cell = 'B' + str(rowselector)
  ## Blank Column for Updating Time Sent
  update_cell = 'D' + str(rowselector)
  ## Blank Column for Updating UUID
  id_cell = 'A' + str(rowselector)
  uniqueid = str(uuid.uuid4())
  sheet[id_cell] = uniqueid
  email = sheet[cur_cell].value
  insert_cal_event(account,email,uniqueid)
  sheet[update_cell] = now.strftime("%m/%d/%Y, %H:%M:%S")
  wb.save(fn)
  rowselector += 1
